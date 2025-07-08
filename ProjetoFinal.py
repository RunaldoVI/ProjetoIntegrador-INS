import nltk, re, numpy as np
import pandas as pd
from pdfminer.high_level import extract_pages, extract_text
from pdfminer.layout import LTTextContainer
from pdf2image import convert_from_path
import pytesseract
from sklearn.multiclass import OneVsRestClassifier
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
import unicodedata

for recurso in ['stopwords', 'punkt', 'wordnet', 'omw-1.4']:
    nltk.download(recurso)
    
# 1. PRE-PROCESSAMENTO
def preprocess(text):
    text = re.sub(r'[^a-zA-Z]', ' ', text.lower())
    tokens = nltk.word_tokenize(text)
    stop_words = set(stopwords.words('english'))
    lem = WordNetLemmatizer()
    return [lem.lemmatize(t) for t in tokens if t not in stop_words]

# 2A. FUNÇÃO ANTIGA PARA PDFs COMO DPQ_J.pdf
def extrair_dpq(pdf_path):
    perguntas = []
    pergunta_atual = ""
    respostas = []

    for page_layout in extract_pages(pdf_path):
        for element in page_layout:
            if isinstance(element, LTTextContainer):
                linhas = element.get_text().strip().splitlines()
                for linha in linhas:
                    linha = linha.strip()
                    if not linha:
                        continue
                    if re.match(r'^DPQ\.\d{3}', linha):
                        if pergunta_atual and respostas:
                            perguntas.append((pergunta_atual, respostas))
                        pergunta_atual = linha
                        respostas = []
                    elif "HANDCARD" in linha or "HAND CARD" in linha:
                        continue
                    elif re.search(r'\bnot at all\b|\bseveral days\b|\bmore than half\b|\bnearly every day\b|\bsomewhat\b|\bvery\b|\bextremely\b|\brefused\b|\bdon’t know\b', linha, re.IGNORECASE):
                        respostas.append(linha)
                    else:
                        if not respostas:
                            pergunta_atual += " " + linha
    if pergunta_atual and respostas:
        perguntas.append((pergunta_atual, respostas))

    return perguntas

# 2B. GENÉRICA PARA OUTROS PDFs
def extrair_frases_unificadas(pdf_path):
    texto = extract_text(pdf_path)
    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    frases = []
    frase_atual = ""

    for i, linha in enumerate(linhas):
        if re.match(r'^\d+\.$', linha) and i + 1 < len(linhas):
            frase_atual = linha + " " + linhas[i + 1]
            continue
        elif frase_atual:
            if linha and not frase_atual.endswith(('.', '?', ':')) and linha[0].islower():
                frase_atual += " " + linha
            else:
                frases.append(frase_atual.strip())
                frase_atual = linha
        else:
            frase_atual = linha

    if frase_atual:
        frases.append(frase_atual.strip())
    return frases

# 3. CARREGAR GLOVE
def load_glove(path):
    emb = {}
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            parts = line.split()
            word, vec = parts[0], np.array(parts[1:], dtype=float)
            emb[word] = vec
    return emb

glove = load_glove('glove/glove.6B.100d.txt')
EMB_SIZE = 100

def sent_to_vec(tokens):
    vecs = [glove[w] for w in tokens if w in glove]
    return np.mean(vecs, axis=0) if vecs else np.zeros(EMB_SIZE)

# 4. TREINO DO MODELO MULTICLASSE
df = pd.read_csv("training_data.csv")
corpus = df['texto'].tolist()
y = df['label'].tolist()

tokens_list = [preprocess(t) for t in corpus]
X = np.vstack([sent_to_vec(t) for t in tokens_list])
X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42)

clf = OneVsRestClassifier(LogisticRegression(max_iter=1000))
y_tr = [str(label) for label in y_tr]
clf.fit(X_tr, y_tr)

# 5. CLASSIFICAÇÃO MULTICLASSE
label_map = {0: "Identificador", 1: "Secção", 2: "Pergunta", 3: "Resposta"}

def classificar_linhas(linhas):
    out = []
    for frase in linhas:
        frase = frase.strip()
        
        # Separar identificador DPQ.xxx
        match = re.match(r'^(DPQ\.\d{3})(.*)', frase)
        if match:
            identificador = match.group(1).strip()
            restante = match.group(2).strip()
            out.append((identificador, "Identificador"))
            
            if restante:
                if ":" in restante:
                    partes = restante.split(":", 1)
                    secao = partes[0].replace("[", "").strip() + ":"
                    pergunta = partes[1].replace("]", "").strip()
                    
                    # Classifica secção
                    tokens_sec = preprocess(secao)
                    vec_sec = sent_to_vec(tokens_sec).reshape(1, -1)
                    pred_sec = int(clf.predict(vec_sec)[0])
                    tipo_sec = label_map.get(pred_sec, "Desconhecido")
                    out.append((secao, tipo_sec))

                    # Classifica pergunta
                    tokens_perg = preprocess(pergunta)
                    vec_perg = sent_to_vec(tokens_perg).reshape(1, -1)
                    pred_perg = int(clf.predict(vec_perg)[0])
                    tipo_perg = label_map.get(pred_perg, "Desconhecido")
                    out.append((pergunta, tipo_perg))
                else:
                    tokens = preprocess(restante)
                    vec = sent_to_vec(tokens).reshape(1, -1)
                    pred = int(clf.predict(vec)[0])
                    tipo = label_map.get(pred, "Desconhecido")
                    out.append((restante, tipo))
        else:
            # Frase sem identificador → classifica como está (pode ser resposta)
            tokens = preprocess(frase)
            vec = sent_to_vec(tokens).reshape(1, -1)
            pred = int(clf.predict(vec)[0])
            tipo = label_map.get(pred, "Desconhecido")
            out.append((frase, tipo))
    return out
# 6. EXECUÇÃO
pdf_path = "DPQ_J.pdf"  # ou "PHQ9.pdf"

if "DPQ" in pdf_path.upper():
    linhas_dpq = extrair_dpq(pdf_path)
    frases_individuais = []
    for pergunta, respostas in linhas_dpq:
        frases_individuais.append(pergunta)
        frases_individuais.extend(respostas)
else:
    frases_individuais = extrair_frases_unificadas(pdf_path)

resultados = classificar_linhas(frases_individuais)

# 6.5. ESCREVER NO EXCEL


def escrever_no_excel(blocos, caminho_excel):
    wb = load_workbook(caminho_excel)
    ws = wb["Detectors"]

    # Obter todos os identificadores já presentes na coluna A
    identificadores_existentes = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            identificadores_existentes.add(str(row[0]).strip())

    linha_inicial = ws.max_row + 1

    for bloco in blocos:
        identificador = bloco.get("Identificador")
        pergunta = bloco.get("Pergunta")

        if identificador and pergunta and identificador not in identificadores_existentes:
            ws.cell(row=linha_inicial, column=1).value = identificador
            ws.cell(row=linha_inicial, column=2).value = "Detector"
            ws.cell(row=linha_inicial, column=4).value = pergunta
            ws.cell(row=linha_inicial, column=5).value = "DETSTEM-GENERIC-DETECTOR"
            linha_inicial += 1

    wb.save(caminho_excel)
    print(f"\n✅ Dados atualizados na folha 'Detectors' de {caminho_excel} com sucesso!")
# Chamar a função com todos os blocos processados:

def escrever_codebooks_no_excel(blocos, caminho_excel):
    wb = load_workbook(caminho_excel)
    if "CodeBooks" not in wb.sheetnames:
        print("❌ A folha 'CodeBooks' não existe no Excel.")
        return

    ws = wb["CodeBooks"]

    # Obter todos os hasURIs já presentes para evitar duplicados
    uris_existentes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # coluna A
            uris_existentes.add(row[0])

    # Encontrar próxima linha disponível
    linha_inicial = 2
    while ws.cell(row=linha_inicial + 1, column=1).value:
        linha_inicial += 1

    for bloco in blocos:
        identificador = bloco.get("Identificador")
        pergunta = bloco.get("Pergunta")
        respostas = bloco.get("Respostas", [])

        if not identificador or not pergunta:
            continue

        # Linha para a pergunta
        uri_pergunta = f"nhaces:CB-{identificador}"
        if uri_pergunta not in uris_existentes:
            ws.cell(row=linha_inicial + 1, column=1).value = uri_pergunta
            ws.cell(row=linha_inicial + 1, column=2).value = "vstoi:Codebook"
            ws.cell(row=linha_inicial + 1, column=3).value = "vstoi:Codebook"
            ws.cell(row=linha_inicial + 1, column=4).value = f"PHQ-9:{pergunta}"
            uris_existentes.add(uri_pergunta)
            linha_inicial += 1

        # Linhas para as respostas
        for resposta in respostas:
            match = re.match(r'^(.*?)(\d+)$', resposta.strip())
            if not match:
                continue
            texto_resp = match.group(1).strip().rstrip(".:;-").title()
            valor_resp = match.group(2).strip()
            uri_resp = f"nhaces:CB-{identificador}-{valor_resp}"

            if uri_resp not in uris_existentes:
                ws.cell(row=linha_inicial + 1, column=1).value = uri_resp
                ws.cell(row=linha_inicial + 1, column=2).value = "vstoi:Codebook"
                ws.cell(row=linha_inicial + 1, column=3).value = "vstoi:Codebook"
                ws.cell(row=linha_inicial + 1, column=4).value = f"{valor_resp} - {texto_resp}"
                uris_existentes.add(uri_resp)
                linha_inicial += 1

    wb.save(caminho_excel)
    print(f"✅ Dados escritos na folha 'CodeBooks' de {caminho_excel} com sucesso!")




#--------------------------------------------------------
def normalizar_resposta(texto):
    texto = texto.lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = re.sub(r'\bor\b', '', texto)
    texto = re.sub(r'[^\w\s]', '', texto)
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto

def escrever_response_options(blocos, caminho_excel, questionario_nome="DPQ"):
    wb = load_workbook(caminho_excel)
    ws = wb["ResponseOptions"]

    # Encontra a primeira linha vazia
    linha = 2
    while ws.cell(row=linha, column=1).value:
        linha += 1

    contador = 0
    respostas_inseridas = set()

    # Lê todas as respostas já escritas no Excel (a partir do label LIMPO)
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[3]  # rdfs:label
        if label:
            chave = normalizar_resposta(label)
            respostas_inseridas.add(chave)

        uri = row[0]
        if uri and uri.startswith(f"nhanes:{questionario_nome},_"):
            match = re.match(rf"nhanes:{questionario_nome},_(\d+)", uri)
            if match:
                contador = max(contador, int(match.group(1)) + 1)

    novas = 0

    for bloco in blocos:
        respostas = bloco.get("Respostas", [])

        novas_respostas = []
        for resposta in respostas:
            label_limpo = resposta.strip()
            
            # Remove padrões como "......", ".....   1", etc
            label_limpo = re.sub(r'[\.\·]{2,}', '', label_limpo)
            
            # Remove o número no final (ex: "   1", " 2", etc)
            label_limpo = re.sub(r'\s*\d+\s*$', '', label_limpo)
            
            # Remove pontuação no final (ex: "?", ".", ",")
            label_limpo = re.sub(r'[.,!?…]+$', '', label_limpo).strip()
            
            # Capitaliza só a primeira letra, o resto em minúsculas
            label_limpo = label_limpo[:1].upper() + label_limpo[1:].lower()
            chave = normalizar_resposta(label_limpo)

            if chave not in respostas_inseridas:
                novas_respostas.append((resposta, label_limpo, chave))

        if not novas_respostas:
            print("⛔ Pergunta ignorada — todas as respostas já estão no Excel.")
            continue

        for resposta, label_limpo, chave in novas_respostas:
            respostas_inseridas.add(chave)

            numero_match = re.search(r'(\d+)\s*$', resposta)
            numero = numero_match.group(1) if numero_match else ""

            ws.cell(row=linha, column=1).value = f"nhanes:{questionario_nome},_{contador}"
            ws.cell(row=linha, column=2).value = "FrequencyResponse"
            ws.cell(row=linha, column=3).value = "Response"
            ws.cell(row=linha, column=4).value = label_limpo
            ws.cell(row=linha, column=5).value = numero
            ws.cell(row=linha, column=6).value = "en"
            ws.cell(row=linha, column=7).value = "1"

            linha += 1
            contador += 1
            novas += 1

    wb.save(caminho_excel)
    print(f"\n✅ Tabela 'ResponseOptions' atualizada com {novas} novas respostas.")
#------------------------------------------------------


def escrever_codebook_slots_corrigido(caminho_excel, questionario_nome="DPQ"):
    wb = load_workbook(caminho_excel)
    ws_slots = wb["CodeBookSlots"]
    ws_respostas = wb["ResponseOptions"]
    ws_codebooks = wb["CodeBooks"]

    # Mapa: (identificador, num_resposta) → URI da resposta (ex: nhanes:DPQ_0)
    resposta_map = {}  # chave = (identificador, num), valor = uri

    for row in ws_respostas.iter_rows(min_row=2, values_only=True):
        uri, _, _, label, num, *_ = row
        if uri and num is not None:
            chave_normalizada = re.sub(r'[^a-zA-Z0-9]', '', label or "").lower()
            resposta_map[(chave_normalizada, str(num))] = uri

    # Obter entradas de perguntas no CodeBooks
    perguntas = []
    for row in ws_codebooks.iter_rows(min_row=2, values_only=True):
        uri, _, _, label = row[:4]
        if uri and label and not re.search(r"-\d+$", uri):  # apenas linhas da pergunta (sem -nº)
            match = re.match(r"nhaces:CB-(DPQ\.\d{3})", uri)
            if match:
                identificador = match.group(1)
                perguntas.append((uri, identificador))

    # Obter respostas associadas (linhas com -n no URI)
    respostas = []
    for row in ws_codebooks.iter_rows(min_row=2, values_only=True):
        uri, _, _, label = row[:4]
        if uri and label and re.search(r"-\d+$", uri):
            match = re.match(r"nhaces:CB-(DPQ\.\d{3})-(\d+)", uri)
            if match:
                identificador = match.group(1)
                numero = match.group(2)
                chave_resp = re.sub(r'[^a-zA-Z0-9]', '', label.split(" - ", 1)[-1]).lower()
                respostas.append((uri, identificador, numero, chave_resp))

    # Verificar quais slots já existem para evitar duplicados
    slots_existentes = set()
    for row in ws_slots.iter_rows(min_row=2, values_only=True):
        if row[0]:
            slots_existentes.add(row[0])

    linha = ws_slots.max_row + 1
    novos_slots = 0

    for uri_resp, identificador, numero, texto_normalizado in respostas:
        key = (texto_normalizado, numero)
        uri_response_option = resposta_map.get(key)
        if not uri_response_option:
            continue

        uri_slot = f"nhanes:CB-{identificador}-{numero}"
        if uri_slot in slots_existentes:
            continue

        uri_codebook = f"nhanes:CB-{identificador}"

        ws_slots.cell(row=linha, column=1).value = uri_slot
        ws_slots.cell(row=linha, column=2).value = "vstoi:CodeBookSlots"
        ws_slots.cell(row=linha, column=3).value = "vstoi:CodeBookSlots"
        ws_slots.cell(row=linha, column=4).value = uri_codebook
        ws_slots.cell(row=linha, column=5).value = uri_response_option

        linha += 1
        novos_slots += 1

    wb.save(caminho_excel)
    print(f"\n✅ Tabela 'CodeBookSlots' atualizada com {novos_slots} novas ligações.")


#--------------------------------------------------------
blocos_processados = []

if resultados:
    bloco_temp = {}
    for frase, tipo in resultados:
        if tipo == "Identificador":
            if bloco_temp:
                blocos_processados.append(bloco_temp)
            bloco_temp = {'Identificador': frase, 'Secção': '', 'Pergunta': '', 'Respostas': []}
        elif tipo == "Secção":
            bloco_temp['Secção'] = frase
        elif tipo == "Pergunta":
            bloco_temp['Pergunta'] = frase
        elif tipo == "Resposta":
            bloco_temp.setdefault('Respostas', []).append(frase)

    # Adiciona o último bloco
    if bloco_temp:
        blocos_processados.append(bloco_temp)
print("\nClassificação estruturada do conteúdo do PDF:\n" + "-"*60)

bloco_atual = {}
for frase, tipo in resultados:
    if tipo == "Identificador":
        if bloco_atual:
            # Mostra bloco anterior antes de iniciar o novo
            print(f"\nIdentificador - {bloco_atual.get('Identificador')}")
            print(f"Secção - {bloco_atual.get('Secção')}")
            print(f"Pergunta - {bloco_atual.get('Pergunta')}")
            for resp in bloco_atual.get('Respostas', []):
                print(f"Resposta- {resp}")
        # Inicia novo bloco
        bloco_atual = {'Identificador': frase, 'Secção': '', 'Pergunta': '', 'Respostas': []}
    elif tipo == "Secção":
        bloco_atual['Secção'] = frase
    elif tipo == "Pergunta":
        bloco_atual['Pergunta'] = frase
    elif tipo == "Resposta":
        bloco_atual.setdefault('Respostas', []).append(frase)

# Imprime último bloco
if bloco_atual:
    print(f"\nIdentificador - {bloco_atual.get('Identificador')}")
    print(f"Secção - {bloco_atual.get('Secção')}")
    print(f"Pergunta - {bloco_atual.get('Pergunta')}")
    for resp in bloco_atual.get('Respostas', []):
        print(f"Resposta - {resp}")
    escrever_no_excel(blocos_processados, "INS-NHANES-2017-2018-DPQ_J.xlsx")
    escrever_codebooks_no_excel(blocos_processados, "INS-NHANES-2017-2018-DPQ_J.xlsx")
    escrever_response_options(blocos_processados, "INS-NHANES-2017-2018-DPQ_J.xlsx", questionario_nome="DPQ")
    escrever_codebook_slots_corrigido("INS-NHANES-2017-2018-DPQ_J.xlsx")

